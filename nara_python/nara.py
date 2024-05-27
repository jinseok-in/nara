import requests 
import re
import math
import os
from bs4 import BeautifulSoup 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from time import sleep
from datetime import datetime
from urllib import parse
from openpyxl.styles import *

#함수
def keyword(val):
    encodingKeyWord = parse.quote(val, encoding="cp949")
    url = rf'https://www.g2b.go.kr:8340/body.do?kwd={encodingKeyWord}&category=TGONG&subCategory=ALL&detailSearch=false&sort=R&reSrchFlag=false&pageNum=1&srchFd=ALL&date=&startDate=&endDate=&startDate2=&endDate2=&orgType=balju&orgName=&orgCode=&swFlag=Y&dateType=&area=&gonggoNo=&preKwd=&preKwds=&body=yes&orgSrchFlag=false'
    return url
#URL 변환

def clean_text(text):
    cleaned_text = re.sub(r'[a-zA-Z]' , '', text)
    cleaned_text = re.sub(r'[\{\}\[\]?.,;|\)*~`!^\-_+<>@\#$%&\\\=\(\'\"]', '', cleaned_text)
    cleaned_text = cleaned_text.replace("\n", "")
    cleaned_text = cleaned_text.replace("\r", "")
    cleaned_text = cleaned_text.replace("\t", "")
    return cleaned_text
#특수문자 필터링 (제어문자)

def sb_mp_title_trans(text):
    cleaned_text = re.sub(r'[\[\]\-]', '', text)
    x_text = cleaned_text[:-2]
    return x_text
#[]- 필터링, 뒤에서 2글자 삭제

def timeTrans(timeList):
    format = r'%Y/%m/%d %H:%M'
    str_datetime = datetime.strptime(timeList, format)
    str_datetime = math.trunc(datetime(
        str_datetime.year, 
        str_datetime.month, 
        str_datetime.day, 
        str_datetime.hour, 
        str_datetime.minute
        ).timestamp())
    return str_datetime
#시간 변환 [0000/00/00 00:00] -> [Epoch Time 0000000000] (math.trunc : 소수점 버림)
# Epoch Time 7days - 604800
# 1 month (30.44 days)

# font
#noteFont = Font(name="", size="", color="")


# 검색할 키워드
keywords = [
    "iso27001",
    "isms",
    "취약점진단",
    "관리체계",
    "수준강화"
    ]
find_excel_keyword = [
    "공고기관",
    "개찰(입찰)일시",
    "개찰일시",
    "배정예산",
    "공고명"
]
timeNow = datetime.today()
startTime = str(timeNow.time())

# 엑셀
wb = Workbook()
ws_Note = wb.active
ws_Note.title = "Note"

ws_Note['B2'] = "Note"
ws_Note['B3'] = "마감기한이 일주일 이상 남은 공고를 가져옵니다."
ws_Note['B4'] = "\"공고 기관\", \"입찰 일시\", \"사업 금액 (추정가격 + 부가세)\", \"공고 명\"이 저장됩니다."

options =  ChromeOptions()
options.add_argument('headless')
options.add_argument("--disable-extensions")  # 확장 프로그램 비활성화
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")  # overcome limited resource problems
options.add_argument('--disable-blink-features=AutomationControlled')

# 드라이버 옵션
print("Start Time : ", timeNow.date(), startTime[:-7])

if not os.path.exists('C:/RFP'):
    os.makedirs('C:/RFP')

if not os.path.exists(f'C:/RFP/{timeNow.date()}'):
    os.makedirs(f'C:/RFP/{timeNow.date()}')

# 크롬 드라이버 최신 버전 설정
service = ChromeService(executable_path=ChromeDriverManager().install())
        

for k in (range(len(keywords))) : # 키워드 반복
    # chrome driver
    driver = webdriver.Chrome(service=service, options=options) # <- options로 변경
    tabs = driver.window_handles
    Data = [requests.get(keyword(keywords[k]))]
    soup = BeautifulSoup(Data[0].content, 'html.parser')

    bid_Deadline = []   # 입찰마감일
    bid_Deadline_list = [] 
    mp_title = []       # 공고 코드 (형식 [코드-차수] [E000000000-00])
    sb_mp_title = ""    # 공고 코드에서 특수문자, 차수 제거
    folderNm_AnAg = []  # 폴더 이름 공고 기관
    folderNm = ""
    search_url = "" # 검색 후 클릭된 url
    epochToday = timeTrans(str(datetime.today()).replace('-', '/')[0:16]) #오늘 날짜 Epochtime 변환
    
    bid_Deadline_list = soup.select('li.m1 > span')
    print("************************\nKeyword start : %s\n************************" % keywords[k])

    #키워드 별 sheet 생성
    ws_keyword = wb.active
    wb.create_sheet(f'{keywords[k]}', k+1)
    ws = wb[f'{keywords[k]}']
    ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'] = '공고 기관', '입찰 일시', '개찰 일시', '배정 예산', '공고 명'

    for i in range(len(bid_Deadline_list)) :
        bid_Deadline.append(timeTrans(clean_text(soup.select('ul.search_list > li > ul.info2 > li.m1 > span')[i].text)))  
        if int(bid_Deadline[i]) > (int(epochToday) - 604800) : # 실행 기준 마감일이 7일 이상 남았을 경우 실행 (7일 604800) 하루 86400
            mp_title.append(soup.select('ul > li > strong > a > span.num')[i].text) # 공고번호 가져오기
        else :
            pass

    for j in range(1, len(mp_title)+1) :
        sb_mp_title = sb_mp_title_trans(mp_title[j-1]) # 공고번호 변환
        search_title = keyword(sb_mp_title) # 공고번호로 검색한 주소
        search_data = [requests.get(search_title)]
        search_soup = BeautifulSoup(search_data[0].content, 'html.parser')
        driver.get(search_title) # 공고번호로 검색된 주소 크롬에서 열기
        
        #셀레니움 요소 찾기
        driver.implicitly_wait(15)
        driver.find_elements(By.XPATH, '/html/body/ul/li/strong/a')[0].click() # 검색된 사이트 클릭
        driver.implicitly_wait(15)
        driver.switch_to.frame('bodyFrame')
        #본 페이지에서 iframe으로 전환
        exNum = j + 2
        ex_cell_li = []
        driver.implicitly_wait(15)

        try :
            driver.find_element(By.XPATH, '//*[@id="epDialogBtns"]/a').click()
            driver.implicitly_wait(15)
        except :
            pass

        # if문, for문을 이용해서 우선 공고기관을 찾고 그 위치에 있는 텍스트를 가져오면 될 것 같음
        # 딕셔너리 구조로 변경?
        # //*[@id="container"]/div[5]/table/tbody/tr[i]/th[1]/p - 공고기관 tr[4]
        if driver.find_element(By.XPATH, '//*[@id="container"]') :
            print(driver.find_elements(By.XPATH, '//*[@id="container"]/div[4]/h3')) # 공고일반 찾기
            #driver_container = driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th/p')
            for i in range(3, len(driver.find_elements(By.XPATH, '//*[@id="container"]/div/table/tbody/tr/th[1]/p'))) :
                if driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th/p')[i].text == '공고기관' : # 공고 기관
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[4]/td[1]/div/span').text))
                else :
                    #pass
                    print(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[4]/td[1]/div/span').text)
                    
                if driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th/p')[i].text == '개찰(입찰)일시' : # 개찰(입찰)일시
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[2]/td[1]/div').text))
                else :
                    #pass
                    print(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[2]/td[1]/div').text)
                
                if driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th/p')[i].text == '개찰일시' : # 개찰일시
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="resultForm"]/div{i}/table/tbody/tr[2]/td[1]/div').text))
                else :
                    #pass
                    print(driver.find_element(By.XPATH, f'//*[@id="resultForm"]/div{i}/table/tbody/tr[2]/td[1]/div').text)
                
                if driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th/p')[i].text == '배정예산' : # 사업 금액
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[2]/td[1]/div').text))
                else :
                    #pass
                    print(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[2]/td[1]/div').text)
                
                if driver.find_elements(By.XPATH, f'//*[@id="container"]/div/table/tbody/tr/th[1]/p')[i].text == '공고명' : # 공고명
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[3]/td/div').text))
                else :
                    #pass
                    print(driver.find_element(By.XPATH, f'//*[@id="container"]/div{i}/table/tbody/tr[3]/td/div').text)
                    
                '''
        elif driver.find_element(By.XPATH, '//*[@id="inForm"]') :
            print('//*[@id="inForm"]/div[2]/h3') # 공고일반 찾기
            for i in range(1, len(driver.find_elements(By.XPATH, '//*[@id="inForm"]/div/table/tbody/tr/th/p'))) :
                if driver.find_elements(By.XPATH, '//*[@id="inForm"]/div/table/tbody/tr/th/p')[i].text == '공고기관' :
                    ex_cell_li.append(str(driver.find_element(By.XPATH, f'//*[@id="inForm"]/div[4]/table/tbody/tr[4]/td[1]/div/a/span').text))
                else :
                    print(driver.find_element(By.XPATH, f'//*[@id="inForm"]/div/table/tbody/tr[4]/td[1]/div/a/span').text)
                if driver.find_elements(By.XPATH, '')[i].text == '개찰일시':
                    print()
        '''
        print(ex_cell_li)
        ws[f'B{exNum}'], ws[f'C{exNum}'], ws[f'D{exNum}'], ws[f'E{exNum}'], ws[f'F{exNum}'] = ex_cell_li[0], ex_cell_li[1], ex_cell_li[2], ex_cell_li[3], ex_cell_li[4]

        #공고기관
        try :
            folderNm_AnAg.append(driver.find_element(By.XPATH, '//*[@id="container"]/div/table/tbody/tr[4]/td[1]/div/span').text)
            #파일이 존재할 경우
        except :
            folderNm_AnAg.append(driver.find_element(By.XPATH, '//*[@id="inForm"]/div[4]/table/tbody/tr[4]/td[1]/div/a/span').text)
            #파일이 존재하지 않을 경우
        #발주명, 발주기관명 가져오기 (파일이 존재하는 페이지와 존재하지 않는 페이지의 형태가 달라 예외 처리)
    driver.quit()

    for j in range(0, len(folderNm_AnAg)):
        if not os.path.exists(f'C:/RFP/{timeNow.date()}/{keywords[k]}'):
            os.makedirs(f'C:/RFP/{timeNow.date()}/{keywords[k]}')
        if not os.path.exists(f'C:/RFP/{timeNow.date()}/{keywords[k]}/{folderNm_AnAg[j]}'):
            os.makedirs(f'C:/RFP/{timeNow.date()}/{keywords[k]}/{folderNm_AnAg[j]}')

        print('Download start : ', folderNm_AnAg[j])
        
        download_path = f"C:\RFP\{timeNow.date()}\{keywords[k]}\{folderNm_AnAg[j]}"
        prefs = {"download.default_directory" : download_path,  
                "profile.default_content_setting_values.automatic_downloads": 1, 
                "download.prompt_for_download": False}
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(service=service, options=options)
        #파일 다운로드 경로 지정

        sb_mp_title = sb_mp_title_trans(mp_title[j]) 
        # 공고번호 변환

        search_title = keyword(sb_mp_title) 
        # 공고번호로 검색한 주소
        search_data = [requests.get(search_title)]
        search_soup = BeautifulSoup(search_data[0].content, 'html.parser')
        
        driver.get(search_title) 
        # 공고번호로 검색된 주소 크롬에서 열기

        driver.implicitly_wait(5)
        driver.find_element(By.XPATH, '/html/body/ul/li/strong/a').click() # 검색된 사이트 클릭
        driver.implicitly_wait(5)
        driver.switch_to.frame('bodyFrame')
        try :
            downloadFile = driver.find_elements(By.CSS_SELECTOR, '#container > div > table > tbody > tr > td.tl > div > a')
            for i in range(len(downloadFile)):
                driver.execute_script("arguments[0].click();", downloadFile[i])
        except :
            print("다운로드 파일 없음")
            pass
        sleep(2)
        
        driver.switch_to.default_content()
        driver.quit()

driver.quit()
timeNow = datetime.today()
endTime = str(timeNow.time())
print("End Time : ", timeNow.date(), endTime[:-7])
print('Successfully completed.')
wb.save(rf'C:\RFP\{timeNow.date()}\bid_announcement.xlsx')
os.startfile('C:/RFP') # 폴더 열기